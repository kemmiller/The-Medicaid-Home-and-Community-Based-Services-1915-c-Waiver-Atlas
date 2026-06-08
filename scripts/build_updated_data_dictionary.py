"""Build an updated 1915(c) waiver data dictionary by merging the original
dictionary with observed values from the final extracted CSV dataset."""

from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DICT_PATH = Path(
    "/Users/vigneshrbabu/Documents/HealthPolicyManagement/MedicaidWaiverExtraction/"
    "dictionary/1915 C Data Dictionary.xlsx"
)
CSV_PATH = Path(
    "/Users/vigneshrbabu/Downloads/1915c_Waiver_Extracted_v2/1915c-waiver-level.csv"
)
OUT_DIR = Path(
    "/Users/vigneshrbabu/Documents/Git/"
    "The-Medicaid-Home-and-Community-Based-Services-1915-c-Waiver-Atlas/docs"
)
OUT_FILE = OUT_DIR / "1915C_Data_Dictionary_Updated.xlsx"

ORIGINAL_COLS = [
    "Variable Name",
    "Indicator",
    "Format",
    "Values",
    "Short Description",
    "Long Description",
]
ADDED_COLS = [
    "In Dataset?",
    "Observed Values (from CSV)",
    "Non-Null Count",
    "Missing %",
    "Inferred Type",
]

HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF")
CHECKBOX_FILL = PatternFill("solid", fgColor="DCEBFB")
RADIO_FILL = PatternFill("solid", fgColor="FFF8CC")
YES_FILL = PatternFill("solid", fgColor="C6EFCE")
NO_FILL = PatternFill("solid", fgColor="F4C7C3")

DATE_RE = re.compile(r"^\s*\d{1,4}[-/]\d{1,2}[-/]\d{1,4}(\s+\d{1,2}:\d{2}(:\d{2})?)?\s*$")
BINARY_TOKENS = {"0", "1", "yes", "no", "true", "false", "y", "n"}


def normalize_format(value) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    s = str(value).strip().lower().replace("-", " ").replace("_", " ")
    s = re.sub(r"\s+", " ", s)
    if "check" in s:
        return "Check Box"
    if "radio" in s:
        return "Radio Button"
    if "text" in s:
        return "Text Box"
    return str(value).strip()


def is_blank(v) -> bool:
    if v is None:
        return True
    if isinstance(v, float) and pd.isna(v):
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


def infer_type(series: pd.Series) -> str:
    vals = [v for v in series.tolist() if not is_blank(v)]
    if not vals:
        return "empty"
    str_vals = [str(v).strip() for v in vals]
    lower = [v.lower() for v in str_vals]
    if all(v in BINARY_TOKENS for v in lower):
        return "binary/checkbox"
    numeric_ok = 0
    for v in str_vals:
        try:
            float(v.replace(",", ""))
            numeric_ok += 1
        except ValueError:
            pass
    if numeric_ok == len(str_vals):
        return "numeric"
    date_ok = sum(1 for v in str_vals if DATE_RE.match(v))
    if date_ok >= max(1, int(0.5 * len(str_vals))):
        return "date"
    return "text"


def summarize_column(series: pd.Series) -> dict:
    total = len(series)
    non_null_mask = series.apply(lambda v: not is_blank(v))
    non_null_count = int(non_null_mask.sum())
    missing_pct = round(100.0 * (total - non_null_count) / total, 2) if total else 0.0
    non_null_vals = series[non_null_mask].tolist()
    seen = []
    seen_set = set()
    for v in non_null_vals:
        s = str(v).strip()
        if s not in seen_set:
            seen_set.add(s)
            seen.append(s)
        if len(seen) >= 50:
            break
    observed = "; ".join(seen)
    return {
        "Observed Values (from CSV)": observed,
        "Non-Null Count": non_null_count,
        "Missing %": missing_pct,
        "Inferred Type": infer_type(series),
    }


def style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 28


def set_widths(ws, widths: dict[str, int]) -> None:
    for col_name, width in widths.items():
        for cell in ws[1]:
            if cell.value == col_name:
                ws.column_dimensions[get_column_letter(cell.column)].width = width
                break


def write_dict_sheet(wb: Workbook, name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(name)
    cols = list(df.columns)
    ws.append(cols)
    fmt_idx = cols.index("Format") + 1 if "Format" in cols else None
    in_idx = cols.index("In Dataset?") + 1 if "In Dataset?" in cols else None

    for row in df.itertuples(index=False):
        ws.append(list(row))
        r = ws.max_row
        fmt_val = ws.cell(row=r, column=fmt_idx).value if fmt_idx else None
        fill = None
        if fmt_val == "Check Box":
            fill = CHECKBOX_FILL
        elif fmt_val == "Radio Button":
            fill = RADIO_FILL
        if fill is not None:
            for c in range(1, len(cols) + 1):
                ws.cell(row=r, column=c).fill = fill
        if in_idx:
            cell = ws.cell(row=r, column=in_idx)
            if cell.value == "Yes":
                cell.fill = YES_FILL
            elif cell.value == "No":
                cell.fill = NO_FILL

    style_header(ws, len(cols))
    widths = {
        "Variable Name": 32,
        "Indicator": 14,
        "Format": 14,
        "Values": 40,
        "Short Description": 45,
        "Long Description": 60,
        "In Dataset?": 12,
        "Observed Values (from CSV)": 60,
        "Non-Null Count": 14,
        "Missing %": 12,
        "Inferred Type": 18,
    }
    set_widths(ws, widths)


def write_csv_only_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("CSV-Only Columns")
    cols = list(df.columns)
    ws.append(cols)
    for row in df.itertuples(index=False):
        ws.append(list(row))
    style_header(ws, len(cols))
    set_widths(
        ws,
        {
            "CSV Column": 32,
            "Observed Values (from CSV)": 60,
            "Non-Null Count": 14,
            "Missing %": 12,
            "Inferred Type": 18,
        },
    )


def write_summary_sheet(wb: Workbook, stats: dict) -> None:
    ws = wb.create_sheet("Summary", 0)
    ws.append(["Metric", "Value"])
    for k, v in stats.items():
        ws.append([k, v])
    style_header(ws, 2)
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 18


COL_ALIASES = {
    "Variable Name": ["variable name", "variable"],
    "Indicator": ["indicator"],
    "Format": ["format"],
    "Values": ["values", "value"],
    "Short Description": ["short description", "short desc"],
    "Long Description": ["long description", "long desc", "description"],
}


def rename_dict_columns(df: pd.DataFrame) -> pd.DataFrame:
    mapping = {}
    for canonical, aliases in COL_ALIASES.items():
        for c in df.columns:
            cl = str(c).strip().lower()
            if any(cl == a or cl.startswith(a + " ") or cl.startswith(a + "(") or a in cl for a in aliases):
                if canonical not in mapping.values():
                    mapping[c] = canonical
                    break
    return df.rename(columns=mapping)


FLIP_TO_RADIO = {
    "waive_1902a",
    "waive_statewideness",
    "numberbenes_limited",
    "phaseinoutschedule",
    "specialHCBS",
}

# Row-level renames applied before merges (codebase naming is canonical).
RENAMES = {
    "spousal_improv_a": "spousal_impov_a",
    "spousal_improv_b": "spousal_impov_b",
    "spousal_improv_c": "spousal_impov_c",
    "loc_eval_insturment": "local_eval_instrument",
}

MERGES = [
    {
        "new_name": "costlimit",
        "drop": [
            "costlimit_excsinstocosts",
            "costlimit_instit",
            "cost_limit_excsinst_costs",
            "cost_limit_instit",
            "cost_limit_lowerinstit",
            "cost_limit_nolimit",
        ],
        "indicator": "Cost limit type used by the waiver (Appendix B-2)",
        "values": (
            "1 = No Cost Limit; "
            "2 = Cost Limit in Excess of Institutional Costs; "
            "3 = Institutional Cost Limit (100% of level of care); "
            "4 = Cost Limit Lower Than Institutional Costs"
        ),
        "short": "Cost limit type (radio).",
        "long": (
            "Merged radio variable for the Appendix B-2 cost limit selection "
            "(svapdxB2_1:elgIclType). Emitted by the PDF acroform extractor "
            "and produced by the HTML/text extractors via the radio-collapse "
            "layer over the split flags cost_limit_nolimit / "
            "cost_limit_excsinst_costs / cost_limit_instit / "
            "cost_limit_lowerinstit. The companion text field "
            "cost_limit_pcntaboveinstit (\"specify the percentage\") remains a "
            "separate Text Box."
        ),
    },
    {
        "new_name": "spousal_impov_bc",
        "drop": ["spousal_improv_b", "spousal_improv_c", "spousal_impov_b", "spousal_impov_c"],
        "indicator": "Post-eligibility rule election when spousal impoverishment rules apply (Appendix B-5)",
        "values": (
            "1 = Use spousal post-eligibility rules; "
            "2 = Use regular post-eligibility rules"
        ),
        "short": "Post-eligibility rule choice (radio).",
        "long": (
            "Merged radio variable from svapdxB5_1:elgIncSpoImpRlsType. "
            "spousal_impov_a remains a separate Check Box indicating whether "
            "spousal impoverishment rules are used at all."
        ),
    },
    {
        "new_name": "local_eval",
        "drop": ["local_eval_a", "local_eval_b", "local_eval_c", "local_eval_d"],
        "indicator": "Responsibility for performing level-of-care evaluations and reevaluations (Appendix B-6-b)",
        "values": (
            "1 = Directly by the Medicaid agency; "
            "2 = By the operating agency specified in Appendix A; "
            "3 = By a government agency under contract with the Medicaid agency; "
            "4 = Other"
        ),
        "short": "Who performs LOC evaluations (radio).",
        "long": (
            "Merged radio variable from svapdxB6_1:elgEvalRespType. Emitted by "
            "the PDF acroform extractor; HTML extractor produces this via the "
            "radio-collapse layer over local_eval_a/b/c/d."
        ),
    },
    {
        "new_name": "local_eval_instrument",
        "drop": [
            "loc_eval_insturment",
            "local_eval_instrument",
            "local_eval_instrument_same",
            "local_eval_instrument_diff",
        ],
        "indicator": "Whether the LOC instrument differs between waiver and institutional care (Appendix B-6-e)",
        "values": (
            "1 = Same instrument used for waiver and institutional level of care; "
            "2 = Different instrument used for waiver vs. institutional level of care"
        ),
        "short": "LOC instrument same vs different (radio).",
        "long": (
            "Merged radio variable from svapdxB6_1:elgEvalLOCInstType. Replaces "
            "the typo'd dictionary entry loc_eval_insturment. Emitted by the "
            "PDF acroform extractor; HTML extractor produces this via the "
            "radio-collapse layer."
        ),
    },
    {
        "new_name": "sd_election",
        "drop": ["sd_election_1", "sd_election_2", "sd_election_3", "sd_delection3"],
        "indicator": "Election of participant direction (Appendix E-1-d)",
        "values": (
            "1 = Waiver supports only individuals who want to direct their services; "
            "2 = Every participant has the opportunity to elect to direct waiver services; "
            "3 = Opportunity to direct some or all services subject to State criteria"
        ),
        "short": "Self-direction election scope (radio).",
        "long": (
            "Merged radio variable from svapdxE1_3:dosElctn. The original "
            "dictionary held sd_election_1 and sd_election_2; option 3 already "
            "exists in the extractor (html_appendix_e_extractor.sd_election_3). "
            "The Appendix E extractor now emits the merged column via the "
            "radio-collapse layer."
        ),
    },
    {
        "new_name": "sd_authority",
        "drop": ["sd_employerauth", "sd_budgetauth", "sd_bothauth"],
        "indicator": "Participant Direction Opportunities (Appendix E-1-b)",
        "values": (
            "1 = Participant: Employer Authority; "
            "2 = Participant: Budget Authority; "
            "3 = Both Authorities"
        ),
        "short": "Participant direction authority type (radio).",
        "long": (
            "Merged radio variable from svapdxE1_2:dosPtcOppType. Replaces the "
            "three split flags sd_employerauth / sd_budgetauth / sd_bothauth in "
            "the dictionary, since the underlying CMS form is a single \"select "
            "one\" question. The Appendix E extractor still computes the split "
            "flags internally and collapses them on output."
        ),
    },
    {
        "new_name": "statecontracts_mcos",
        "drop": [
            "statecontracts_mcos1",
            "statecontracts_mcos2",
            "statecontracts_mcos3",
            "statecontracts_mcos4",
        ],
        "indicator": "Managed care contracting arrangement for waiver services (Appendix I-3-7)",
        "values": (
            "1 = State does not contract with MCOs/PIHPs/PAHPs; "
            "2 = State contracts with MCO(s)/PIHP(s)/PAHP(s); "
            "3 = Concurrent 1915(b)/1915(c) waiver; "
            "4 = Concurrent 1115/1915(c) waiver; "
            "5 = Uses more than one of the above contract authorities"
        ),
        "short": "MCO/PIHP/PAHP contracting (radio).",
        "long": (
            "Merged radio variable from svapdxI3_7:fnaPymtPHP. Original "
            "dictionary had statecontracts_mcos1-4; the extractor supports 5 "
            "options across template versions."
        ),
    },
]


# Brand-new variables that do not exist in the source dictionary yet (no
# split-flag predecessors to merge from). Appended at the end of the
# dictionary if the new_name is not already present.
NEW_ENTRIES = [
    {
        "new_name":  "min_numservices",
        "indicator": "Minimum number of waiver services required for an individual to be determined to need waiver services (Appendix B-6-a-i)",
        "format":    "Text Box",
        "values":    "Numeric (one or more)",
        "short":     "Minimum number of services to qualify (text box).",
        "long": (
            "Free-text numeric entry from svapdxB6_1:elgEvalSvcMinQty. The state "
            "specifies the minimum count of waiver services an individual must "
            "require in the service plan in order to be determined to need waiver "
            "services. Almost always '1' in the corpus."
        ),
    },
    {
        "new_name":  "reeval_sched",
        "indicator": "Frequency of level-of-care reevaluations (Appendix B-6-g)",
        "format":    "Radio Button",
        "values": (
            "1 = Every three months; "
            "2 = Every six months; "
            "3 = Every twelve months; "
            "4 = Other schedule"
        ),
        "short": "Reevaluation schedule (radio).",
        "long": (
            "Merged radio variable from svapdxB6_1:elgRevalSchType. Emitted by "
            "the PDF acroform extractor; HTML and text extractors produce this "
            "via the radio-collapse layer over reeval_sched_3mo/_6mo/_12mo/_other."
        ),
    },
]


def apply_dictionary_revisions(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    rename_lower = {k.lower(): v for k, v in RENAMES.items()}
    def _rename(v):
        if isinstance(v, str):
            return rename_lower.get(v.strip().lower(), v)
        return v
    df["Variable Name"] = df["Variable Name"].apply(_rename)

    name_lower = df["Variable Name"].astype(str).str.strip().str.lower()
    df.loc[name_lower.isin({n.lower() for n in FLIP_TO_RADIO}), "Format"] = "Radio Button"

    for merge in MERGES:
        drop_lower = {d.lower() for d in merge["drop"]}
        mask = df["Variable Name"].astype(str).str.strip().str.lower().isin(drop_lower)
        if not mask.any():
            continue
        first_idx = df.index[mask].min()
        df = df[~mask].reset_index(drop=True)
        first_idx = min(first_idx, len(df))
        new_row = {
            "Variable Name": merge["new_name"],
            "Indicator": merge["indicator"],
            "Format": "Radio Button",
            "Values": merge["values"],
            "Short Description": merge["short"],
            "Long Description": merge["long"],
        }
        upper = df.iloc[:first_idx]
        lower = df.iloc[first_idx:]
        df = pd.concat(
            [upper, pd.DataFrame([new_row]), lower], ignore_index=True
        )

    name_lower_series = df["Variable Name"].astype(str).str.strip().str.lower()
    for entry in NEW_ENTRIES:
        nm = entry["new_name"].lower()
        mask = name_lower_series == nm
        if mask.any():
            # Row already exists (e.g. min_numservices in the All Variables
            # sheet). Enrich blank cells but leave any user-set content alone.
            idx = df.index[mask][0]
            for src_key, col in (
                ("indicator", "Indicator"),
                ("format",    "Format"),
                ("values",    "Values"),
                ("short",     "Short Description"),
                ("long",      "Long Description"),
            ):
                cur = df.at[idx, col]
                if is_blank(cur):
                    df.at[idx, col] = entry[src_key]
            continue
        new_row = {
            "Variable Name":     entry["new_name"],
            "Indicator":         entry["indicator"],
            "Format":            entry["format"],
            "Values":            entry["values"],
            "Short Description": entry["short"],
            "Long Description":  entry["long"],
        }
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        name_lower_series = df["Variable Name"].astype(str).str.strip().str.lower()

    return df


def process_dict_sheet(
    df_dict: pd.DataFrame, csv_df: pd.DataFrame, csv_lower_map: dict[str, str]
) -> pd.DataFrame:
    df_dict = rename_dict_columns(df_dict)
    for col in ORIGINAL_COLS:
        if col not in df_dict.columns:
            df_dict[col] = ""
    df_dict = df_dict[ORIGINAL_COLS].copy()
    df_dict["Format"] = df_dict["Format"].apply(normalize_format)
    df_dict = apply_dictionary_revisions(df_dict)

    in_dataset, observed, nn_count, miss_pct, inf_type = [], [], [], [], []
    for var in df_dict["Variable Name"].tolist():
        key = str(var).strip().lower() if not is_blank(var) else ""
        actual = csv_lower_map.get(key)
        if actual is None:
            in_dataset.append("No")
            observed.append("")
            nn_count.append(0)
            miss_pct.append(100.0)
            inf_type.append("")
        else:
            in_dataset.append("Yes")
            summary = summarize_column(csv_df[actual])
            observed.append(summary["Observed Values (from CSV)"])
            nn_count.append(summary["Non-Null Count"])
            miss_pct.append(summary["Missing %"])
            inf_type.append(summary["Inferred Type"])

    df_dict["In Dataset?"] = in_dataset
    df_dict["Observed Values (from CSV)"] = observed
    df_dict["Non-Null Count"] = nn_count
    df_dict["Missing %"] = miss_pct
    df_dict["Inferred Type"] = inf_type
    return df_dict


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    all_vars = pd.read_excel(DICT_PATH, sheet_name="All Variables", dtype=object)
    waiver_data = pd.read_excel(DICT_PATH, sheet_name="Waiver Data", dtype=object)
    csv_df = pd.read_csv(CSV_PATH, dtype=object, low_memory=False)

    csv_lower_map = {c.strip().lower(): c for c in csv_df.columns}

    all_vars_out = process_dict_sheet(all_vars, csv_df, csv_lower_map)
    waiver_out = process_dict_sheet(waiver_data, csv_df, csv_lower_map)

    dict_names_lower = set()
    for df in (all_vars_out, waiver_out):
        for v in df["Variable Name"].tolist():
            if not is_blank(v):
                dict_names_lower.add(str(v).strip().lower())

    csv_only_rows = []
    for col in csv_df.columns:
        if col.strip().lower() not in dict_names_lower:
            s = summarize_column(csv_df[col])
            csv_only_rows.append(
                {
                    "CSV Column": col,
                    "Observed Values (from CSV)": s["Observed Values (from CSV)"],
                    "Non-Null Count": s["Non-Null Count"],
                    "Missing %": s["Missing %"],
                    "Inferred Type": s["Inferred Type"],
                }
            )
    csv_only_df = pd.DataFrame(
        csv_only_rows,
        columns=[
            "CSV Column",
            "Observed Values (from CSV)",
            "Non-Null Count",
            "Missing %",
            "Inferred Type",
        ],
    )

    combined_formats = pd.concat(
        [all_vars_out["Format"], waiver_out["Format"]], ignore_index=True
    )
    combined_in = pd.concat(
        [all_vars_out["In Dataset?"], waiver_out["In Dataset?"]], ignore_index=True
    )
    total_vars = len(all_vars_out) + len(waiver_out)
    stats = {
        "Total dictionary variables (All Variables + Waiver Data)": total_vars,
        "  All Variables sheet rows": len(all_vars_out),
        "  Waiver Data sheet rows": len(waiver_out),
        "Check Box variables": int((combined_formats == "Check Box").sum()),
        "Radio Button variables": int((combined_formats == "Radio Button").sum()),
        "Text Box variables": int((combined_formats == "Text Box").sum()),
        "Variables present in CSV": int((combined_in == "Yes").sum()),
        "Variables absent from CSV": int((combined_in == "No").sum()),
        "Total CSV rows": len(csv_df),
        "Total CSV columns": len(csv_df.columns),
        "CSV-only columns (not in dictionary)": len(csv_only_df),
    }

    wb = Workbook()
    wb.remove(wb.active)
    write_dict_sheet(wb, "All Variables", all_vars_out)
    write_dict_sheet(wb, "Waiver Data", waiver_out)
    write_csv_only_sheet(wb, csv_only_df)
    write_summary_sheet(wb, stats)
    wb.save(OUT_FILE)

    print(f"Wrote: {OUT_FILE}")
    print("Summary:")
    for k, v in stats.items():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
